"""
Export helpers: PDF (ReportLab), Excel (OpenPyXL), CSV.

All functions write to the configured export directory and return the absolute
file path. CSV uses the stdlib ``csv`` module (no third-party dependency).
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from app.config import EXPORT_DIR


def _safe_filename(name: str) -> str:
    keep = (" ", ".", "_", "-")
    return "".join(c if c.isalnum() or c in keep else "_" for c in name).strip()


def _export_path(prefix: str, extension: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{prefix}_{stamp}.{extension}"
    return str(EXPORT_DIR / _safe_filename(fname))


def export_csv(rows: Sequence[dict], columns: Sequence[str], prefix: str) -> str:
    """Write rows to a CSV file and return its path."""
    path = _export_path(prefix, "csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in columns})
    return path


def export_excel(rows: Sequence[dict], columns: Sequence[str], prefix: str, sheet_name: str = "Report") -> str:
    """Write rows to an .xlsx file and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = _export_path(prefix, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            ws.cell(row=r_idx, column=col_idx, value=value)

    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(path)
    return path


def export_pdf(
    title: str,
    columns: Sequence[str],
    rows: Sequence[dict],
    prefix: str,
    summary: Iterable[tuple[str, str]] | None = None,
) -> str:
    """Render a simple styled PDF report and return its path."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        Paragraph,
    )

    path = _export_path(prefix, "pdf")
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=18 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleX", parent=styles["Title"], textColor=colors.HexColor("#1E3A8A")
    )
    elements = [
        Paragraph(title, title_style),
        Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 8 * mm),
    ]

    if summary:
        summ_data = [[k, v] for k, v in summary]
        summ_table = Table(summ_data, colWidths=[60 * mm, 100 * mm])
        summ_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EFF6FF")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFDBFE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(summ_table)
        elements.append(Spacer(1, 6 * mm))

    # Table
    header = [c.replace("_", " ").title() for c in columns]
    data = [header]
    for row in rows:
        data.append([str(row.get(c, "")) for c in columns])

    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    return path
