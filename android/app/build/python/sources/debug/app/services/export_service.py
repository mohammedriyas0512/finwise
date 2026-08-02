"""
Export helpers: PDF (ReportLab), Excel (OpenPyXL), CSV.

All functions write to the configured export directory and return the absolute
file path. CSV uses the stdlib ``csv`` module (no third-party dependency).
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from app.config import EXPORT_DIR


def _safe_filename(name: str) -> str:
    keep = (" ", ".", "_", "-")
    return "".join(c if c.isalnum() or c in keep else "_" for c in name).strip()


def _export_path(prefix: str, extension: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{prefix}_{stamp}.{extension}"
    return str(EXPORT_DIR / _safe_filename(fname))


def export_csv(rows: Sequence[dict], columns: Sequence[str], prefix: str) -> str:
    """Write rows to a CSV file and return its path."""
    path = _export_path(prefix, "csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in columns})
    return path


def export_excel(rows: Sequence[dict], columns: Sequence[str], prefix: str, sheet_name: str = "Report") -> str:
    """Write rows to an .xlsx file and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = _export_path(prefix, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            ws.cell(row=r_idx, column=col_idx, value=value)

    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(path)
    return path


def export_pdf(
    title: str,
    columns: Sequence[str],
    rows: Sequence[dict],
    prefix: str,
    summary: Iterable[tuple[str, str]] | None = None,
) -> str:
    """Render a simple styled PDF report (fpdf2) and return its path."""
    from fpdf import FPDF
    from fpdf.fonts import FontFace

    path = _export_path(prefix, "pdf")
    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 18, 15)

    pdf.add_page()
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(30, 58, 138)
    pdf.cell(0, 10, text=title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 6, text=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    if summary:
        pdf.set_font("helvetica", "B", 10)
        for k, v in summary:
            pdf.set_font("helvetica", "B", 10)
            pdf.set_fill_color(239, 246, 255)
            pdf.cell(55, 7, text=str(k), fill=True)
            pdf.set_font("helvetica", "", 10)
            pdf.cell(0, 7, text=str(v), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

    headings_style = FontFace(emphasis="BOLD", color=(255, 255, 255), fill_color=(37, 99, 235))
    with pdf.table(headings_style=headings_style, text_align="LEFT",
                   width=180, line_height=5) as table:
        header = table.row()
        for c in columns:
            header.cell(c.replace("_", " ").title())
        for row in rows:
            r = table.row()
            for c in columns:
                r.cell(str(row.get(c, "")))

    pdf.output(path)
    return path
